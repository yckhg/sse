import io
import unittest

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from odoo.tests import HttpCase


@unittest.skipIf(load_workbook is None, "openpyxl not available")
class TestAssetTemplate(HttpCase):

    def test_download_asset_template(self):
        self.authenticate('admin', 'admin')
        company = self.env.company
        account_model = self.env['account.account'].with_company(company)

        # Archive all existing accounts to control the test environment
        account_model.search([]).write({'active': False})

        account_model.create({
            'name': 'Fixed Asset Account',
            'code': 'FA001',
            'account_type': 'asset_non_current',
        })
        account_model.create({
            'name': 'Expense Account',
            'code': 'EX001',
            'account_type': 'expense',
        })

        self.env['account.journal'].with_company(company).create({
            'name': 'Miscellaneous Operations',
            'code': 'MISC-test',
            'type': 'general',
        })

        response = self.url_open(f'/web/binary/download_asset_template/{company.id}')
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook['Assets']

        expected_rows = [
            ('Computer 1', '800.00', '01-01-2025', '', 'Straight Line', '4', 'Years', '', 'No Prorata', '01-01-2025',
             '200.00', '0.00', company.name, 'FA001 Fixed Asset Account', 'FA001 Fixed Asset Account', 'EX001 Expense Account', 'Miscellaneous Operations'),
            ('Computer 2', '25000.00', '03-15-2025', '', 'Declining', '5', 'Years', '2', 'Based on days per period', '03-15-2025',
             '0.00', '2000.00', company.name, 'FA001 Fixed Asset Account', 'FA001 Fixed Asset Account', 'EX001 Expense Account', 'Miscellaneous Operations'),
            ('Machine A', '15000.00', '09-01-2024', '', 'Straight Line', '10', 'Years', '', 'Constant Periods', '09-01-2024',
             '1500.00', '500.00', company.name, 'FA001 Fixed Asset Account', 'FA001 Fixed Asset Account', 'EX001 Expense Account', 'Miscellaneous Operations'),
            ('Machine B', '100000.00', '06-20-2025', '', 'Straight Line', '60', 'Months', '', 'No Prorata', '06-20-2025',
             '10000.00', '10000.00', company.name, 'FA001 Fixed Asset Account', 'FA001 Fixed Asset Account', 'EX001 Expense Account', 'Miscellaneous Operations'),
        ]

        actual_rows = [
            tuple('' if v is None else str(v) for v in actual_row[:17])
            for actual_row in sheet.iter_rows(min_row=2, values_only=True)
        ]
        self.assertEqual(actual_rows, expected_rows)
