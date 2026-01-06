import contextlib
import datetime
import json
import logging
import math
import re
import requests

from io import BytesIO
from lxml import html
from openpyxl import load_workbook
from requests.exceptions import HTTPError, RequestException

from odoo import Command, api, fields, models
from odoo.exceptions import ValidationError
from odoo.tools import format_date
from odoo.tools.float_utils import float_compare

_logger = logging.getLogger(__name__)


class EsgDatabase(models.Model):
    _name = 'esg.database'
    _description = 'Database'

    name = fields.Char(required=True)
    url = fields.Char()
    image = fields.Binary()
    last_update = fields.Date()
    latest_version = fields.Date()
    color = fields.Integer(compute='_compute_kanban_display')
    kanban_text = fields.Char(compute='_compute_kanban_display')
    can_be_downloaded = fields.Boolean(compute='_compute_can_be_downloaded')

    @api.depends('latest_version', 'last_update')
    def _compute_kanban_display(self):
        for db in self:
            db.color = False
            db.kanban_text = self.env._("Not loaded")
            if not db.last_update:
                continue
            formatted_date = format_date(self.env, db.last_update, date_format='MMMM y')
            db.kanban_text = self.env._("Updated on: %(date)s", date=formatted_date)
            if db.last_update >= db.latest_version:
                db.color = 7
            else:
                db.color = 2

    def _compute_can_be_downloaded(self):
        ademe_db = self.env.ref('esg.esg_database_ademe')
        ipcc_db = self.env.ref('esg.esg_database_ipcc')
        for database in self:
            if database in (ademe_db, ipcc_db):
                database.can_be_downloaded = True
            else:
                database.can_be_downloaded = False

    @api.ondelete(at_uninstall=False)
    def _prevent_database_deletion(self):
        ademe_db = self.env.ref('esg.esg_database_ademe')
        ipcc_db = self.env.ref('esg.esg_database_ipcc')
        if ademe_db in self:
            raise ValidationError(self.env._("You can't delete the ADEME database."))
        if ipcc_db in self:
            raise ValidationError(self.env._("You can't delete the IPCC database."))

    def action_load_data(self):
        self.ensure_one()
        if self == self.env.ref('esg.esg_database_ademe'):
            result = self._action_import_ademe_file()
        elif self == self.env.ref('esg.esg_database_ipcc'):
            result = self._action_import_efdb_from_ipcc()
        else:
            raise ValidationError(self.env._("Database file is missing"))
        if result is True:
            self.last_update = self.latest_version
        action_view_emission_factor = self.env['ir.actions.actions']._for_xml_id('esg.action_view_emission_factor')
        action_view_emission_factor['domain'] = [('database_id', '=', self.id)]
        return action_view_emission_factor

    def _external_api_call(self, request_url):
        ademe_api_url = 'https://data.ademe.fr/data-fair/api/v1/datasets/base-carboner'
        if not request_url.startswith(ademe_api_url):
            raise ValidationError(self.env._('Invalid URL.'))
        response_error = {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'warning',
                },
            }
        try:
            request_response = requests.request(
                'GET',
                request_url,
                timeout=(30, 30),
            )
            request_response.raise_for_status()
        except (ValueError, HTTPError, RequestException) as exception:
            response_error['params']['message'] = self.env._(
                "Server returned an unexpected error: %(error)s",
                error=(request_response.text or str(exception)),
            )
            return response_error

        try:
            response_text = request_response.text
            response_data = json.loads(response_text)
        except json.JSONDecodeError:
            response_error['params']['message'] = self.env._("JSON response could not be decoded.")
            return response_error

        return response_data

    # =============
    # ADEME IMPORT
    # =============
    def _get_ademe_emission_factor_values(self, lines):
        existing_factors = dict(self.env['esg.emission.factor']._read_group(
            domain=[('database_id', '=', self.id)],
            groupby=['code'],
            aggregates=['id:recordset'],
        ))
        gasses = {
            'CO2f': self.env.ref('esg.esg_gas_co2'),
            'CH4f': self.env.ref('esg.esg_gas_ch4f'),
            'CH4b': self.env.ref('esg.esg_gas_ch4b'),
            'N2O': self.env.ref('esg.esg_gas_n2o'),
            'SF6': self.env.ref('esg.esg_gas_sf6'),
        }
        hardcoded_units_conversion = {
            'kg': [
                'kg of active substance', 'kg (live weight)', 'kg net weight', 'kg of spreaded nitrogen',
                'kg of suppressed DCO', 'kg of treated leather', 'Kg', 'kg of spreaded nitrgen',
                'kgH2', 'kg NTK', 'kg BioGNC'
            ],
            'L': ['liter', 'Liter'],
            'Ton': ['ton', 'ton of K2O', 'ton of N', 'ton of P2O5', 'ton of clinker', 'ton of waste', 'tonne'],
            'Units': ['unit', 'meal'],
            'm': ['m of road'],
            'm²': ['m²  net floor area', 'm² of ceiling', 'm² of floor', 'm² of wall', 'm2 SHON'],
            'm³': ['m3', 'm3 (n)'],
            'ml': ['mL'],
            'Hours': ['hour', 'heure'],
            'kWh': ['kWh (LHV)', 'kWh (PCI)', 'kWh HHV', 'kWh ICV', 'kWh IVC', 'kWh LHV', 'kWh SCV']
        }
        hardcoded_source_scopes = {
            'Achats de biens': 'indirect_others',
            'Process et émissions fugitives': 'direct',
            'Achats de services': 'indirect_others',
            'Combustibles': 'direct',
            'Transport de marchandises': 'indirect_others',
            'Transport de personnes': 'indirect_others',
            'Statistiques territoriales': 'direct',
            'UTCF': 'direct',
            'Traitement des déchets': 'indirect_others',
            'Electricité': 'indirect',
            'Réseaux de chaleur / froid': 'direct',
        }

        def _get_gas_lines(line, gasses, total_emissions):
            gas_lines = []
            for gas in ['CO2f', 'CH4f', 'CH4b', 'N2O']:
                if line.get(gas, 0) > 0:
                    gas_lines.append({
                        'gas_id': gasses[gas],
                        'quantity': line[gas],
                        'activity_type_id': line.get('Type_poste', False),
                    })
            for i in range(1, 6):
                additional_gas = line.get(f"Code_gaz_supplémentaire_{i}", None)
                if additional_gas and additional_gas in ['SF6']:
                    gas_lines.append({
                        'gas_id': gasses[additional_gas],
                        'quantity': line.get(f"Valeur_gaz_supplémentaire_{i}", 0),
                        'activity_type_id': line.get('Type_poste', False),
                    })
            # if the gaz decomposition is equal to the total emissions, that means that the decomposition
            # is given with CO2 equivalent value; each line needs to be adapted to have the gaz volume instead
            if not float_compare(sum(gas['quantity'] for gas in gas_lines), total_emissions, precision_rounding=0.1):
                for gas in gas_lines:
                    gas['quantity'] = gas['quantity'] / gas['gas_id'].global_warming_potential

            return gas_lines

        _logger.info("ESG: Processing raw lines")
        line_values = []
        posts_gas_lines = []
        for raw_line in lines:
            if not raw_line:
                continue
            line = raw_line  # dict(zip(dict_keys, raw_line))
            if line["Statut_de_l'élément"] == "Archivé" or\
                line["Type_de_l'élément"] == "Données source":
                continue

            total_emissions = line['Total_poste_non_décomposé']
            if line["Type_Ligne"] == "Poste":
                posts_gas_lines += _get_gas_lines(line, gasses, total_emissions)
                line_values[-1]['gas_line_ids'] = posts_gas_lines
                continue

            dates = {}
            for raw_date in ['Date_de_création', 'Période_de_validité']:
                if raw_date not in line:
                    dates[raw_date] = False
                    continue
                try:
                    dates[raw_date] = datetime.datetime.strptime(line[raw_date], '%Y-%m-%d')
                except ValueError:
                    dates[raw_date] = False

            posts_gas_lines = []
            gas_lines = _get_gas_lines(line, gasses, total_emissions)

            normalized_unit = line.get('Unité_anglais', line.get('Unité_français', '')).replace('kgCO2e/', '')
            for base_unit, alternate_names in hardcoded_units_conversion.items():
                if normalized_unit in alternate_names:
                    normalized_unit = base_unit
                    break

            factor_values = {
                'name': f'{line.get("Nom_base_anglais", "")} {line.get("Nom_attribut_anglais", "")} {line.get("Nom_frontière_anglais", "")}'
                    if line.get("Nom_base_anglais", '')
                    else f'{line.get("Nom_base_français", "")} {line.get("Nom_attribut_français", "")} {line.get("Nom_frontière_français", "")}',
                'code': line["Identifiant_de_l'élément"],
                'database_id': self.id,
                'uom_id': normalized_unit,
                'esg_uncertainty_value': line.get('Incertitude', 0) / 100,
                'compute_method': 'physically',
                'valid_from': dates['Date_de_création'],
                'valid_to': dates['Période_de_validité'],
                'source_id': line['Code_de_la_catégorie'],
                'gas_line_ids': gas_lines,
            }
            if not gas_lines:
                factor_values['esg_emissions_value'] = total_emissions
            line_values.append(factor_values)

        _logger.info("ESG: Checking existing units, emission sources and esg activity type")
        # Prepare to process all units at once to avoid multiplying queries and loops
        units = [line['uom_id'].strip() for line in line_values]
        existing_units = dict(
            self.env['uom.uom']._read_group(
                domain=[('name', 'in', units)],
                groupby=['name'],
                aggregates=['id:recordset'],
            )
        )
        # Prepare to process all sources at once to avoid multiplying queries and loops
        source_tree_list = []
        for line in line_values:
            source_tree_list.append([s.strip() for s in line['source_id'].split('>')])
        all_sources_name = {item for sublist in source_tree_list for item in sublist}
        existing_sources = {(name, parent): record for name, parent, record in self.env['esg.emission.source']._read_group(
            domain=[('name', 'in', all_sources_name)],
            groupby=['name', 'parent_id'],
            aggregates=['id:recordset'],
        )}

        # Prepare to process all activity types at once to avoid multiplying queries and loops
        existing_activity_types = dict(
            self.env['esg.activity.type']._read_group(
                domain=[],
                groupby=['name'],
                aggregates=['id:recordset'],
            )
        )

        _logger.info("ESG: Processing link to related records")
        # slow loop, should be optimized
        for source_tree, unit, line in zip(source_tree_list, units, line_values):
            # units
            existing_unit = existing_units.get(unit, False)
            if not existing_unit:
                existing_unit = self.env['uom.uom'].create({'name': unit})
                existing_units[unit] = existing_unit
            line['uom_id'] = existing_unit.id
            # sources
            previous_existing_source = self.env['esg.emission.source']
            for source_name in source_tree:
                existing_source = existing_sources.get((source_name, previous_existing_source), False)
                if not existing_source:
                    existing_source = self.env['esg.emission.source'].create({
                        'name': source_name,
                        'parent_id': previous_existing_source.id,
                        'scope': hardcoded_source_scopes.get(source_name, previous_existing_source.scope) or 'direct',
                    })
                    existing_sources[source_name, previous_existing_source] = existing_source
                previous_existing_source = existing_source
            line['source_id'] = previous_existing_source.id
            # activity types + declare gas lines correctly
            if line['gas_line_ids']:
                for gas_line in line['gas_line_ids']:
                    gas_line['gas_id'] = gas_line['gas_id'].id
                    if not gas_line['activity_type_id']:
                        continue
                    activity = existing_activity_types.get(gas_line['activity_type_id'], False)
                    if not activity:
                        activity = self.env['esg.activity.type'].create({'name': gas_line['activity_type_id']})
                        existing_activity_types[gas_line['activity_type_id']] = activity
                    gas_line['activity_type_id'] = activity.id
                line['gas_line_ids'] = [Command.create(gas) for gas in line['gas_line_ids']]

        new_factor_values = []
        write_factor_values = {}
        for line_value in line_values:
            # Separate data into existing one and new one
            line_code = line_value['code']
            if line_code in existing_factors:
                write_factor_values[existing_factors[line_code]] = line_value
            else:
                new_factor_values.append(line_value)
        return new_factor_values, write_factor_values

    def _action_import_ademe_file(self):
        ademe_api_url = 'https://data.ademe.fr/data-fair/api/v1/datasets/base-carboner'
        ademe_data = self._external_api_call(ademe_api_url)
        if not ademe_data.get('updatedAt'):
            return ademe_data
        ademe_data['updatedAt'] = re.sub(r'\.\d+', '', ademe_data['updatedAt'])
        self.latest_version = datetime.datetime.strptime(ademe_data['updatedAt'], "%Y-%m-%dT%H:%M:%SZ")
        if self.last_update and self.latest_version <= self.last_update:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'success',
                    'message': self.env._('ADEME data is already up to date.'),
                },
            }

        required_calls = math.ceil(ademe_data['count'] / 10000)  # the api limits the amount of data lines to 10k

        _logger.info("ESG: Fetching ADEME data from their API")
        complete_data = []
        for i in range(0, required_calls):
            data_url = f'{ademe_api_url}/lines?size=10000&format=json&after={10000 * i}'
            data = self._external_api_call(data_url)
            if not data.get('results', None):
                return data
            complete_data += data['results']

        _logger.info("ESG: Reading ADEME carbon data")
        try:
            new_values, write_values = self._get_ademe_emission_factor_values(complete_data)
            _logger.info("ESG: Creating/Writing records")
            self.env['esg.emission.factor'].create(new_values)
            for record, values in write_values.items():
                record.gas_line_ids = False
                record.write(values)
            _logger.info("ESG: File imported")
        except KeyError as e:
            _logger.error(e)
            raise ValidationError(self.env._("The file format doesn't seem to be correct."))
        return True

    # =============
    # IPCC IMPORT
    # =============
    def _get_ipcc_xls_file(self, reset=False):
        request_headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Content-Type': 'application/x-www-form-urlencoded',
        }
        table_name = 'tmp_e63ckbntq05n963ul2451jk2sp'
        if reset:
            request_response = requests.request(
                'GET',
                'https://www.ipcc-nggip.iges.or.jp/EFDB/find_ef.php?reset=',
                timeout=(30, 30),
            )
            request_response.raise_for_status()
            request_response = requests.request(
                'POST',
                'https://www.ipcc-nggip.iges.or.jp/EFDB/find_ef.php',
                headers=request_headers,
                data={'action': 'apply_filter', 'source_data': 'default'},
                timeout=(30, 30),
            )
            request_response.raise_for_status()
            root_node = html.fromstring(request_response.content)
            if elements := root_node.xpath('//input[@name="tableName"]'):
                if len(elements) == 1:
                    table_name = elements[0].value
            cookie = request_response.headers.get('Set-Cookie', '')
            cookie_parts = [part.split('=', 1) for part in cookie.split(';') if '=' in part]
            for cookie_name, cookie_value in cookie_parts:
                if cookie_name == 'PHPSESSID':
                    request_headers['Cookie'] = f'PHPSESSID={cookie_value}'
                    break
        request_response = requests.request(
            'POST',
            'https://www.ipcc-nggip.iges.or.jp/EFDB/find_ef_xls.php',
            headers=request_headers,
            data={'lang_id': 1, 'tableName': table_name, 'mi_show_fuel': True, 'mi_show_cpool': True},
            timeout=(30, 30),
        )
        request_response.raise_for_status()
        if request_response.content:
            xls_file = BytesIO(request_response.content)
        elif not reset:
            xls_file = self._get_ipcc_xls_file(reset=True)
        else:
            raise ValidationError(self.env._("The IPCC server did not return any file to import the data."))
        return xls_file

    def _action_import_efdb_from_ipcc(self):
        response_error = {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'type': 'warning',
            },
        }
        ipcc_ef_data_list = []
        parent_source_names = set()
        try:
            filename = self._get_ipcc_xls_file()
            workbook = load_workbook(filename=filename)
            sheet = workbook.active
            column_id_per_vals_key = {}

            for row in sheet.iter_rows():
                ipcc_data = {}
                is_header = True
                for cell in row:
                    if cell.row == 1:
                        if cell.value == 'EF ID':
                            column_id_per_vals_key[cell.column] = 'code'
                        elif cell.value == 'IPCC 1996 Source/Sink Category':
                            column_id_per_vals_key[cell.column] = 'parent_source_id'
                        elif cell.value == 'IPCC 2006 Source/Sink Category':
                            column_id_per_vals_key[cell.column] = 'source_id'
                        elif cell.value == 'Gas':
                            column_id_per_vals_key[cell.column] = 'gas_id'
                        elif cell.value == 'Description':
                            column_id_per_vals_key[cell.column] = 'name'
                        elif cell.value == 'Region / Regional Conditions':
                            column_id_per_vals_key[cell.column] = 'region'
                        elif cell.value == 'Value':
                            column_id_per_vals_key[cell.column] = 'quantity'
                        elif cell.value == 'Unit':
                            column_id_per_vals_key[cell.column] = 'unit'
                        elif cell.value == 'Technologies / Practices':
                            column_id_per_vals_key[cell.column] = 'note1'
                        elif cell.value == 'Parameters / Conditions':
                            column_id_per_vals_key[cell.column] = 'note2'
                    elif cell.column in column_id_per_vals_key:
                        is_header = False
                        key = column_id_per_vals_key[cell.column]
                        value = str(cell.value).strip()
                        if key == 'parent_source_id':
                            parent_source_names.add(value)
                        ipcc_data[key] = value
                if not is_header:
                    ipcc_ef_data_list.append(ipcc_data)
        except (ValueError, HTTPError, RequestException) as exception:
            response_error['params']['message'] = self.env._(
                "Server returned an unexpected error: %(error)s",
                error=str(exception),
            )
            return response_error

        parent_sources = self.env['esg.emission.source']._load_records([{'xml_id': f'esg.ipcc_emission_source_{i}', 'noupdate': True, 'values': {'name': source_name.strip()}} for i, source_name in enumerate(parent_source_names, 1)])
        parent_source_per_name = {es.name: es for es in parent_sources}
        source_name_per_parent_source = {
            ef_data['source_id']: parent_source_per_name[ef_data['parent_source_id'].strip()]
            for ef_data in ipcc_ef_data_list
            if ef_data.get('source_id') and parent_source_per_name.get(ef_data.get('parent_source_id', '').strip())
        }
        source_per_name = {
            es.name: es
            for es in self.env['esg.emission.source']._load_records([
                {'xml_id': f'esg.ipcc_emission_source_{parent.id}_{i}', 'noupdate': True, 'values': {'name': source_name.strip(), 'parent_id': parent.id}}
                for i, (source_name, parent) in enumerate(source_name_per_parent_source.items(), 1)
            ])
        }
        gas_mapping = {
            'METHANE': self.env.ref('esg.esg_gas_ch4'),
            'CARBON DIOXIDE': self.env.ref('esg.esg_gas_co2'),
            'c-C4F8': self.env.ref('esg.esg_gas_pfc_c-c4f8'),
            'C2F6': self.env.ref('esg.esg_gas_pfc_c2f6'),
            'C3F8': self.env.ref('esg.esg_gas_pfc_c3f8'),
            'C4F6': self.env.ref('esg.esg_gas_pfc_c4f6'),
            'C4F8O': self.env.ref('esg.esg_gas_c4f8o'),
            'C5F8': self.env.ref('esg.esg_gas_pfc_c5f8'),
            'C6F14': self.env.ref('esg.esg_gas_pfc_c6f14'),
            'CARBON MONOXIDE': self.env.ref('esg.esg_gas_co'),
            'CF4': self.env.ref('esg.esg_gas_pfc_cf4'),
            'HFC-125': self.env.ref('esg.esg_gas_hfc_125'),
            'HFC-134a': self.env.ref('esg.esg_gas_hfc_134a'),
            'HFC-134a\nHFC-152a': self.env.ref('esg.esg_gas_hfc_134a'),
            'HFC-143a': self.env.ref('esg.esg_gas_hfc_143a'),
            'HFC-152a': self.env.ref('esg.esg_gas_hfc_152a'),
            'HFC-23': self.env.ref('esg.esg_gas_hfc_23'),
            'HFC-23\nHFC-32\nHFC-125\nHFC-134a\nHFC-143a\nCF4\nC2F6\nC3F8\nC4F8\nC5F8\nC6F14': self.env.ref('esg.esg_gas_hfc_23'),
            'HFC-23\nHFC-32\nHFC-125\nHFC-134a\nHFC-152a\nHFC-143a\nHFC-227ea\nHFC-236fa': self.env.ref('esg.esg_gas_hfc_23'),
            'HFC-23\nHFC-32\nHFC-41\nHFC-43-10mee\nHFC-125\nHFC-134\nHFC-134a\nHFC-152a\nHFC-143\nHFC-143a\nHFC-227ea\nHFC-236fa\nHFC-245ca\nCF4\nC2F6\nC3F8\nC4F10\nc-C4F8\nC5F12\nC6F14': self.env.ref('esg.esg_gas_hfc_23'),
            'HFC-23\nHFC-32\nHFC-41\nHFC-43-10mee\nHFC-125\nHFC-134\nHFC-134a\nHFC-152a\nHFC-143\nHFC-143a\nHFC-227ea\nHFC-236fa\nHFC-245ca\nHFC-152\nHFC-161\nHFC-236cb\nHFC-236ea\nHFC-245fa\nHFC-365mfc': self.env.ref('esg.esg_gas_hfc_23'),
            'HFC-32': self.env.ref('esg.esg_gas_hfc_32'),
            'HFC-41': self.env.ref('esg.esg_gas_hfc_41'),
            'HFE-125\nHFC-43-10mee\nHFC-125\nHFC-134\nHFC-134a\nHFC-152a\nHFC-143\nHFC-143a\nHFC-227ea\nHFC-236fa\nHFC-245ca\nHFC-152\nHFC-161\nHFC-236cb\nHFC-236ea\nHFC-245fa\nHFC-365mfc': self.env.ref('esg.esg_gas_cf3ochf2'),
            'HFE-245fa1\nHFE-365mcf3\nHFC-134a\nHFC-152a\nHFC-227ea': self.env.ref('esg.esg_gas_chf2ch2ocf3'),
            'HFE-245fa1\nHFE-365mcf3\nHFC-43-10mee\nHFC-134a\nHFC-152a\nHFC-227ea': self.env.ref('esg.esg_gas_chf2ch2ocf3'),
            'HFE-365mcf3\nHFC-43-10mee\nC6F14': self.env.ref('esg.esg_gas_cf3cf2ch2och3'),
            'HFE-7100': self.env.ref('esg.esg_gas_chf2_c4f9och3'),
            'METHANE\nCARBON DIOXIDE\nNITROUS OXIDE': self.env.ref('esg.esg_gas_ch4'),
            'METHANE\nNITROUS OXIDE': self.env.ref('esg.esg_gas_ch4'),
            "NITROGEN OXIDES (NO+NO2)\nMETHANE\nCARBON MONOXIDE\nCARBON DIOXIDE\nNITROUS OXIDE": self.env.ref('esg.esg_gas_co'),
            "NITROGEN OXIDES (NO+NO2)\nMETHANE\nCARBON MONOXIDE\nNITROUS OXIDE": self.env.ref('esg.esg_gas_co'),
            'Nitrogen Trifluoride': self.env.ref('esg.esg_gas_nf3'),
            "Nitrogen Trifluoride\nHFC-23\nHFC-32\nCF4\nC2F6\nC3F8\nc-C4F8\nSulphur Hexafluoride": self.env.ref('esg.esg_gas_nf3'),
            'NITROUS OXIDE': self.env.ref('esg.esg_gas_n2o'),
            "SULPHUR DIOXIDE (SO2+SO3)\nNITROGEN OXIDES (NO+NO2)\nNON METHANE VOLATILE ORGANIC COMPOUNDS\nMETHANE\nCARBON MONOXIDE\nCARBON DIOXIDE\nNITROUS OXIDE": self.env.ref('esg.esg_gas_so2'),
            "Sulphur Hexafluoride": self.env.ref('esg.esg_gas_sf6'),
        }

        def parse_float(value):
            amount = None
            with contextlib.suppress(ValueError, TypeError):
                amount = float(value)
            return amount

        emission_factor_xmlid_list = []
        uom_kg = self.env.ref('uom.product_uom_kgm')
        uom_unit = self.env.ref('uom.product_uom_unit')
        uom_m3 = self.env.ref('uom.product_uom_cubic_meter')
        uom_tonne = self.env.ref('uom.product_uom_ton')
        uom_ha, uom_lto = self.env['uom.uom']._load_records([
            {'xml_id': 'esg.uom_ha', 'noupdate': True, 'values': {'name': 'ha', 'relative_factor': 1}},
            {'xml_id': 'esg.uom_lt', 'noupdate': True, 'values': {'name': 'LTO', 'relative_factor': 1}},
        ])
        nb_skipped_records = 0
        for ef_data in ipcc_ef_data_list:
            code = ef_data['code']
            note = ef_data.get('note1', '')
            uom = uom_kg
            gaz_id = False
            source_id = False
            name = ef_data['name'].strip()
            if not name:
                id = int(code)
                if 327476 <= id <= 327568:
                    name = "Combustion factor (Cf) for fires in vegetation types"
                elif 327569 <= id <= 327644:
                    name = "Below-ground biomass (BGB): root-to-shoot ratio"
                elif 327645 <= id <= 328060:
                    name = "Above-ground biomass (AGB): net biomass growth in natural forests"
                elif 327427 <= id <= 327475:
                    name = "Soil organic carbon stocks (SOCREF) in mineral soils"
                elif 327386 <= id <= 327426:
                    name = "Dead wood carbon stock"
                elif 327260 <= id <= 327385:
                    name = "Litter carbon stock"
                else:
                    nb_skipped_records += 1
                    continue
            if source_name := ef_data['source_id'].strip():
                source = source_per_name.get(source_name)
                if not source:
                    nb_skipped_records += 1
                    continue
                source_id = source.id
            else:
                parent_source = parent_source_per_name.get(ef_data['parent_source_id'].strip())
                if not parent_source:
                    nb_skipped_records += 1
                    continue
                source_id = parent_source.id
            if gaz := gas_mapping.get(ef_data['gas_id']):
                gaz_id = gaz.id
            else:
                nb_skipped_records += 1
                continue
            if note2 := ef_data.get('note2', ''):
                note += '\n' + note2
            value = parse_float(ef_data['quantity'])
            if value is None:
                nb_skipped_records += 1
                continue
            if unit := ef_data.get('unit'):
                if unit.startswith('%') or unit.lower().startswith('fraction') or unit.lower().startswith('year') or unit in ['per year', 'months', 'parts per billion by volume', 'asse', 'installe', 'equipment']:
                    nb_skipped_records += 1
                    continue
                elif unit.startswith('g'):
                    value /= 1000
                elif (
                    unit.lower().startswith('kg')
                ):
                    value *= 1
                elif unit.startswith('TJ'):
                    value *= 1.11 * (10 ** -5)
                elif unit.lower().startswith('gg'):
                    value *= 10 ** 9
                elif unit.lower().startswith('ton') or unit in ['(kg PFC/tAl)/(AE-Minutes/cellday)', '(kg PFC/tAl)/(mV/day)', 'kg SF6/tonnes magnesium produced or smelted']:
                    uom = uom_tonne
                elif unit == 'm3/m3 beer':
                    value *= 1.020
                    uom = uom_m3
                elif unit == 'm3/m3 ethanol':
                    value *= 789
                    uom = uom_m3
                elif unit == 'kg CH4/head/yr':
                    uom = uom_unit
                elif unit == 't dm/ha':
                    uom = uom_ha
                    value *= 1000
                elif unit == 'kg/LTO':
                    uom = uom_lto
                note += '\nUnit converted in kg: ' + unit
            emission_factor_xmlid_list.append({
                'xml_id': f'esg.ipcc_emission_factor_{code}',
                'noupdate': True,
                'values': {
                    'code': code,
                    'name': name,
                    'description': note,
                    'uom_id': uom.id,
                    'source_id': source_id,
                    'region': ef_data['region'],
                    'database_id': self.id,
                    'gas_line_ids': [
                        Command.create({
                            'gas_id': gaz_id,
                            'quantity': value,
                        }),
                    ],
                },
            })
        if nb_skipped_records:
            _logger.warning("%s entries from IPCC Database were skipped because of missing information", nb_skipped_records)
        self.env['esg.emission.factor']._load_records(emission_factor_xmlid_list)
        return True
