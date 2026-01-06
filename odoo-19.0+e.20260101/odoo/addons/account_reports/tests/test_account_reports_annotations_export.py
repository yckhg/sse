import datetime
import io
import unittest

import markupsafe

from odoo.tests import freeze_time, tagged
from odoo.tools import html2plaintext

from odoo.addons.account_reports.tests.common import TestAccountReportsCommon

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@tagged('post_install', '-at_install')
class TestAccountReportAnnotationsExport(TestAccountReportsCommon):
    @classmethod
    def setUpClass(cls):
        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not available")

        super().setUpClass()

        cls.report = cls.env.ref('account_reports.balance_sheet')
        cls.report.column_ids.sortable = True

        # Get accounts
        bank_default_account = cls.company_data["default_journal_bank"].default_account_id
        tax_sale_default_account = cls.company_data["default_account_tax_sale"]
        # Create move
        move = cls.env['account.move'].create({
            'move_type': 'entry',
            'date': '2024-06-10',
            'journal_id': cls.company_data['default_journal_cash'].id,
            'line_ids': [
                (0, 0, {'debit':  250.0,     'credit':   0.0,    'account_id': bank_default_account.id}),
                (0, 0, {'debit':   0.0,     'credit': 250.0,    'account_id': tax_sale_default_account.id}),
            ],
        })
        move.action_post()
        # Create annotation
        with freeze_time('2024-06-20'):
            cls.env.cr._now = datetime.datetime.now()  # used to force create_date, as sql is not wrapped by freeze gun
            date = datetime.datetime.strptime('2024-06-20', '%Y-%m-%d').date()
            message = cls.env['mail.message'].create({
                'model': bank_default_account._name,
                'res_id': bank_default_account.id,
                'body': 'Papa a vu le fifi de lolo',
                'date': date,
                'author_id': cls.env.user.partner_id.id,
                'message_type': 'comment',
                'subtype_id': cls.env.ref('mail.mt_note').id,
            })
            cls.env['account.report.annotation'].create({
                'date': date,
                'message_id': message.id,
            })

    def read_xlsx_data(self, report_data):
        report_file = io.BytesIO(report_data)
        xlsx = load_workbook(filename=report_file, data_only=True)
        sheet = xlsx.worksheets[0]
        return list(sheet.values)

    def test_annotations_export_no_comparison(self):
        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True})
        report_data = self.report.export_to_xlsx(options)
        export_content = self.read_xlsx_data(report_data['file_content'])

        # When there is no comparison, there are two columns for the accounts (number, name), one column for the data and
        # finally one column for the annotations. Hence the index 3 for the annotation column.
        self.assertEqual(export_content[0][3], "Annotations")
        self.assertEqual(export_content[7][3], "1 - Papa a vu le fifi de lolo")

    def test_annotations_export_same_period_last_year_comparison(self):
        default_comparison = {
            'filter': 'same_last_year',
            'number_period': 1,
            'date_from': '2023-06-01',
            'date_to': '2023-06-30',
            'periods': [{
                'string': 'As of 06/30/2023',
                'period_type': 'month', 'mode':
                'single', 'date_from': '2023-06-01',
                'date_to': '2023-06-30'
            }],
            'period_order':
            'descending',
            'string': 'As of 06/30/2023',
            'period_type': 'month',
            'mode': 'single'
        }
        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True, 'comparison': default_comparison})
        report_data = self.report.export_to_xlsx(options)
        export_content = self.read_xlsx_data(report_data['file_content'])

        # When comparing with the same period last year, there are two columns for the accounts (number, name), two columns for the
        # date (one per period), one column for the growth comparison percentage and finally one column for the annotations.
        # Hence the index 5 for the annotation column.
        self.assertEqual(export_content[0][5], "Annotations")
        self.assertEqual(export_content[7][5], "1 - Papa a vu le fifi de lolo")

    def test_annotations_export_two_last_periods_comparison(self):
        default_comparison = {
            'filter': 'previous_period',
            'number_period': 2,
            'date_from': '2024-05-01',
            'date_to': '2024-05-31',
            'periods': [
                {
                    'string': 'As of 05/31/2024',
                    'period_type': 'month',
                    'mode': 'single',
                    'date_from': '2024-05-01',
                    'date_to': '2024-05-31'
                },
                {
                    'string': 'As of 04/30/2024',
                    'period_type': 'month',
                    'mode': 'single',
                    'date_from': '2024-04-01',
                    'date_to': '2024-04-30'
                }
            ],
            'period_order': 'descending',
            'string': 'As of 05/31/2024',
            'period_type': 'month',
            'mode': 'single'
        }
        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True, 'comparison': default_comparison})
        report_data = self.report.export_to_xlsx(options)
        export_content = self.read_xlsx_data(report_data['file_content'])

        # When comparing with the two last periods, there are two columns for the accounts (number, name), three columns for the
        # data (one per period) and finally one column for the annotations. Hence the index 5 for the annotation column.
        self.assertEqual(export_content[0][5], "Annotations")
        self.assertEqual(export_content[7][5], "1 - Papa a vu le fifi de lolo")

    def test_annotations_ordering(self):
        bank_default_account = self.company_data["default_journal_bank"].default_account_id

        model = bank_default_account._name
        res_id = bank_default_account.id
        with freeze_time('2024-06-21'):
            self.env.cr._now = datetime.datetime.now()  # used to force create_date, as sql is not wrapped by freeze gun
            date = datetime.datetime.strptime('2024-06-21', '%Y-%m-%d').date()
            message = self.env['mail.message'].create({
                'model': model,
                'res_id': res_id,
                'body': 'lolo a vu le fifi de papa',
                'date': date,
                'author_id': self.env.user.partner_id.id,
                'message_type': 'comment',
                'subtype_id': self.env.ref('mail.mt_note').id,
            })
            self.env['account.report.annotation'].create({
                'date': date,
                'message_id': message.id,
            })

        options = self._generate_options(self.report, '2024-06-01', '2024-06-30', default_options={'unfold_all': True, 'export_mode': True})
        report_data = self.report.get_report_information(options)
        annotations = report_data['annotations']

        report_line_ids = [f'~account.report.line~{self.report.line_ids[i].id}' for i in range(0, 3)]
        expected = {
            f'~account.report~{self.report.id}|{"|".join(report_line_ids)}|{{"groupby": "account_id"}}~{model}~{res_id}': [
                {
                    'model': model,
                    'res_id': res_id,
                    'date': datetime.date(2024, 6, 20),
                    'body': 'Papa a vu le fifi de lolo',
                },
                {
                    'id': message.id,
                    'model': model,
                    'res_id': res_id,
                    'date': datetime.date(2024, 6, 21),
                    'body': 'lolo a vu le fifi de papa',
                }
            ]
        }
        self.assertTrue(len(annotations.keys()) == len(expected.keys()), "More/Less lines are annotated than expected.")
        self.assertTrue(all(key in annotations for key in expected), "The same lines are not annotated as expected.")
        for line_id, line_annotations in expected.items():
            self.assertEqual(len(annotations[line_id]), len(line_annotations), f"Line {line_id} has more/less annotations than expected.")
            for i in range(len(line_annotations)):
                for field in line_annotations[i]:
                    field_data = annotations[line_id][i][field]
                    if field == 'body':
                        field_data = html2plaintext(markupsafe.escape(field_data))
                    self.assertEqual(field_data, line_annotations[i][field], f"Line {line_id} annotation {i} has a different {field} than expected.")
