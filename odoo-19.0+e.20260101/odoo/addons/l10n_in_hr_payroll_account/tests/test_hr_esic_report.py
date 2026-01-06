import io
import base64
import unittest

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from datetime import date

from odoo.addons.l10n_in_hr_payroll_account.tests.common import TestPayrollAccountCommon
from odoo.tests import tagged


@tagged('post_install_l10n', 'post_install', '-at_install')
class TestHrESICReport(TestPayrollAccountCommon):

    def test_hr_esic_report(self):
        """ Check reminder activity is set the for probation contract
        Test Case
        ---------
            1) Create payslips for bot employee in the contract period
            2) Generate the esic report with the export xlxs button
            3) Check the generated xlxs report has the expected values or not
        """
        payslips = self.env['hr.payslip'].create([{
                'name': 'Jethalal Payslip',
                'employee_id': self.jethalal_emp.id,
                'version_id': self.contract_jethalal.id,
                'date_from': date(2023, 1, 1),
                'date_to': date(2023, 1, 31),
            }, {
                'name': 'Rahul Payslip',
                'employee_id': self.rahul_emp.id,
                'version_id': self.contract_rahul.id,
                'date_from': date(2023, 1, 1),
                'date_to': date(2023, 1, 31),
            }
        ])

        payslips.compute_sheet()
        payslips.action_payslip_done()
        gross = payslips[0]._get_line_values(['GROSS'])['GROSS'][payslips[0].id]['total']
        working_days = sum(workday.number_of_days for workday in payslips[0].worked_days_line_ids if workday.is_paid)

        esic_report = self.env['l10n.in.hr.payroll.esic.report'].create({
            'month': '1',
            'year': '2023',
            'export_report_type': 'esic',
            'xlsx_filename': 'esic'
        })

        # Action to generate the XLSX file
        esic_report.action_export_xlsx()
        self.assertTrue(esic_report.xlsx_file, "The XLS file was not generated.")

        xlsx_data = base64.b64decode(esic_report.xlsx_file)

        if load_workbook is None:
            raise unittest.SkipTest("openpyxl not available")
        xlsx = load_workbook(io.BytesIO(xlsx_data))
        sheet = xlsx.worksheets[0]
        sheet_values = list(sheet.values)

        expected_values = {
            # Header
            0: [
                "IP Number (10 Digits)",
                "IP Name (Only alphabets and space)",
                "No of Days for which wages paid/payable during the month",
                "Total Monthly Wages",
                "Reason Code for Zero workings days(Numeric Only: provide 0 for all other reasons)",
                "Last Working Day (Format DD/MM/YYYY or DD-MM-YYYY)",
            ],
            # Rows
            1: [
                '93487475100284657',
                'Jethalal',
                working_days,
                gross,
                0.0,
                '31-01-2023',
            ],
            2: [
                '93874944361284657',
                'Rahul',
                working_days,
                gross,
                0.0,
                '31-01-2023',
            ],
        }
        for row, values in expected_values.items():
            for row_value, expected_value in zip(sheet_values[row], values):
                self.assertEqual(row_value, expected_value)
