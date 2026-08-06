#!/usr/bin/env python3
"""U1 — 원장 엑셀을 결정론적 정규화 데이터셋으로 변환한다.

계약: docs/spec/greenpm-data-migration/u1-normalized-extract.md
호스트에서 실행하며 Odoo/DB에 접근하지 않는다. 파일 → 파일 변환만 한다.
"""
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent.parent
SRC = BASE / "raw" / "그린피엠_매출 원부자재 외주 관리_260703.xlsx"
OUT = BASE / "migration" / "normalized.json"

# 원장 매출 시트의 카테고리 컬럼 위치 → 정규화 카테고리명
SALES_CATEGORIES = {6: "일반 재활용", 7: "현수막", 8: "배너", 9: "가로등", 10: "어깨띠", 11: "기타"}

# 연도가 생략된 외주 시기의 확정값 (계약 "연도가 생략된 시기의 확정" 표)
MONTH_ONLY_YEAR = {
    ("외주1", "디자인", "10월"): "2024-10-01",
    ("DZ원", "실사", "11월"): "2024-11-01",
    ("DZ원", "실사", "12월"): "2024-12-01",
    ("세무법인", "관리비", "11월"): "2024-11-01",
    ("세무법인", "관리비", "12월"): "2024-12-01",
    **{("세무법인", "관리비", f"{m}월"): f"2025-{m:02d}-01" for m in range(1, 10)},
}


def d(v):
    return v.strftime("%Y-%m-%d") if isinstance(v, datetime) else None


def i(v):
    return int(round(v)) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def s(v):
    return v.strip() if isinstance(v, str) and v.strip() else None


def normalize_sales(ws):
    """매출관리 — 헤더 5행 아래가 사실 행. 거래처가 비면 서식행."""
    out = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not s(row[2]):
            continue
        qty = {name: i(row[col]) for col, name in SALES_CATEGORIES.items() if i(row[col])}
        out.append({
            "일자": d(row[1]),
            "거래처": s(row[2]),
            "담당자": s(row[3]),
            "작업명": s(row[4]),
            "수량": qty,
            "총계": i(row[12]) or 0,
            "총액": i(row[16]),
            "결제완료": bool(row[17]),
            "발송일": d(row[20]),
            "발송방식": s(row[21]),
            "받는곳": s(row[22]),
            "결제방식": s(row[15]),
            "계산서상태": s(row[23]),
        })
    return out


def normalize_materials(ws):
    """원부자재 — '원자재'/'부자재' 구역 마커로 자재구분을 정하고, 구역 헤더 반복행은 버린다.
    '* 앤위브 …' 주석 아래는 사용현황 구역."""
    rows = list(ws.iter_rows(values_only=True))
    items, usage = [], []
    section = None
    in_usage = False
    for row in rows:
        c1, c2 = s(row[1]), s(row[2])
        if c1 and c1.startswith("*"):          # 사용현황 구역 시작 주석
            in_usage = True
            continue
        if c1 in ("원자재", "부자재"):           # 구역 마커
            section = c1
            continue
        if c1 in ("원부자재", "구분"):            # 제목행 / 구역 헤더 반복행
            continue
        if in_usage:
            if c1 == "제작물" or not c1:
                continue
            usage.append({"제작물": c1, "사용일": d(row[2]), "사용량": s(row[3])})
            continue
        if not c1 or not c2:
            continue
        items.append({
            "구분": c1,
            "자재구분": section,
            "매입처": c2,
            "매입가": i(row[3]),
            "매입일": d(row[6]),
            "매입량": i(row[7]),
            "점검일": d(row[8]),
            "잔여량": i(row[9]),
            "스토리지": s(row[10]),
            "결제완료": bool(row[4]),
        })
    return items, usage


def normalize_outsourcing(ws):
    """외주정산 — 매입처는 있으나 구분이 빈 행은 소계행이므로 제외한다."""
    out, subtotals = [], {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        vendor = s(row[1])
        if not vendor or vendor == "매입처":
            continue
        kind = s(row[2])
        if not kind:                                  # 소계행
            subtotals[vendor] = i(row[5])
            continue
        when = d(row[3])
        if when is None:                              # 월만 적힌 시기 → 확정 연월
            key = (vendor, kind, s(row[3]))
            if key not in MONTH_ONLY_YEAR:
                sys.exit(f"확정 연월 미정의: {key}")
            when = MONTH_ONLY_YEAR[key]
        out.append({
            "매입처": vendor,
            "구분": kind,
            "시기": when,
            "정산시기": s(row[4]) or (d(row[4]) if row[4] else None),
            "금액": i(row[5]),
            "결제완료": bool(row[6]),
            "증빙": s(row[7]),
            "비고": s(row[9]),
        })
    return out, subtotals


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    sales = normalize_sales(wb["매출관리"])
    materials, usage = normalize_materials(wb["원부자재"])
    outsourcing, subtotals = normalize_outsourcing(wb["외주정산"])

    data = {
        "매출": sales,
        "원부자재": materials,
        "원단사용현황": usage,
        "외주": outsourcing,
        "_외주소계": subtotals,
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"매출 {len(sales)} · 원부자재 {len(materials)} · 사용현황 {len(usage)} · 외주 {len(outsourcing)}")
    print(f"→ {OUT}")


if __name__ == "__main__":
    main()
